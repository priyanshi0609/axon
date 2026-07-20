"""
Document Processing Agent core logic — routes each file type to the
right extractor, falling back to OCR for scanned/image-only PDF pages.
"""

import os
import io
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
from docx import Document as DocxDocument
from openpyxl import load_workbook

if os.name == "nt":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


def ocr_image_bytes(image_bytes: bytes) -> str:
    image = Image.open(io.BytesIO(image_bytes))
    return pytesseract.image_to_string(image).strip()


def extract_from_pdf(file_path: str) -> dict:
    doc = fitz.open(file_path)
    full_text = []
    ocr_used = False

    for page in doc:
        page_text = page.get_text().strip()
        if page_text:
            full_text.append(page_text)
        else:
            ocr_used = True
            pix = page.get_pixmap(dpi=200)
            full_text.append(ocr_image_bytes(pix.tobytes("png")))

    doc.close()
    return {"text": "\n\n".join(full_text), "pages": len(full_text), "ocr_used": ocr_used}


def extract_from_image(file_path: str) -> dict:
    with open(file_path, "rb") as f:
        text = ocr_image_bytes(f.read())
    return {"text": text, "pages": 1, "ocr_used": True}


def extract_from_docx(file_path: str) -> dict:
    doc = DocxDocument(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return {"text": "\n".join(paragraphs), "pages": None, "ocr_used": False}


def extract_from_xlsx(file_path: str) -> dict:
    wb = load_workbook(file_path, data_only=True)
    all_text = []
    for sheet in wb.worksheets:
        all_text.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                all_text.append(row_text)
    return {"text": "\n".join(all_text), "pages": len(wb.worksheets), "ocr_used": False}


def extract_text(file_path: str) -> dict:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        return extract_from_pdf(file_path)
    elif ext in {".png", ".jpg", ".jpeg"}:
        return extract_from_image(file_path)
    elif ext == ".docx":
        return extract_from_docx(file_path)
    elif ext in {".xlsx", ".xls"}:
        return extract_from_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type for extraction: {ext}")
